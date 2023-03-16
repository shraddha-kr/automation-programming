import openpyxl

def open_and_read_excel_file(sheetname, file):    
    
    wrkbk = openpyxl.load_workbook(file)
    sh = wrkbk.active
    sales_repr = []
 
    # for row in sh.iter_rows(min_row=1, min_col=1, max_row=1, max_col=4):
    for index, row in enumerate(sh.iter_rows(min_row=2, min_col=1, max_col=4)):
        row_str=""                
        for cell in row:        
            if(cell.value is not None):
                row_str += str(cell.value)+" "
        if(len(row_str)>1):
            one_sales_rep = (row_str[:-1].split(" "))
            print(one_sales_rep)
            sales_repr.append(one_sales_rep)

    return sales_repr

# if __name__ =="__main__":
#     list1 = open_and_read_excel_file("data", "./robocorp/SalesData.xlsx")
#     print(len(list1))
#     for i in list1:        
#         print(i)
#         print(type(i))
